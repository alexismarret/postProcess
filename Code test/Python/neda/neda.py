#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Mar  6 20:19:48 2022

@author: alexis
"""

#===================================
#script for H2O2.xlsx
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

params={'axes.titlesize' : 8, 'axes.labelsize' : 8, 'lines.linewidth' : 0.5,
        'lines.markersize' : 1.5, 'xtick.labelsize' : 3, 'ytick.labelsize' : 5,
        'font.size': 8,'legend.fontsize': 8, 'legend.handlelength' : 1.5,
        'legend.borderpad' : 0.1,'legend.labelspacing' : 0.1, 'axes.linewidth' : 1,
        'figure.autolayout': True}
plt.rcParams.update(params)

#===================================
#generate arrays from data
def extract_data(worksheet,row,col,N):
    
    data = np.empty(N, dtype='object')
    i=0
    for r in worksheet.iter_rows(min_row=row, max_row=N+row-1,
                                 min_col=col, max_col=col,
                                 values_only=True):
        data[i] = r[0]
        i+=1
        
    return (data)

#===================================
#extract data from excel file
def get_data(filename,worksheet,N,row,cols):
    
    #load excel file
    ws = load_workbook(filename)[worksheet]
    
    ID   = extract_data(ws,row,cols[0],N)
    init = extract_data(ws,row,cols[1],N)
    mid  = extract_data(ws,row,cols[2],N)
    end  = extract_data(ws,row,cols[3],N)

    return (ID,init,mid,end)

#===================================
#number of genes
N = 17741

#min value of ratio
min_value = 5.

#===================================
ID_H2O2,init_H2O2,mid_H2O2,end_H2O2 = get_data("H2O2.xlsx","20141114 h2o2 all fpkm.txt (2)",N,row=7,cols=[1,4,5,6])
ID_oxi ,init_oxi ,mid_oxi ,end_oxi  = get_data("Oxidative_Stress.xlsx","Table S1 (2)"      ,N,row=9,cols=[2,3,6,9])

#filter out if 0
cond = ((init_H2O2!=0) & (init_oxi!=0))

ratio1_H2O2 = mid_H2O2[cond] / init_H2O2[cond]
ratio1_oxi  = mid_oxi [cond] / init_oxi [cond]

ratio2_H2O2 = end_H2O2[cond] / init_H2O2[cond]
ratio2_oxi  = end_oxi [cond] / init_oxi [cond]

gtr  = ((ratio1_H2O2>min_value) & (ratio1_oxi>min_value))

ID = ID_H2O2[cond][gtr]


#===================================
#plots
plt.close("all")
fig, ((sub1,sub2)) = plt.subplots(1,2,figsize=(4.1,2),dpi=300,sharex=True)
plt.gcf().subplots_adjust(bottom=0.2)

sub1.axhline(min_value,color="gray",linestyle="dashed")
sub2.axhline(min_value,color="gray",linestyle="dashed")

ax = np.linspace(0,len(ID)-1,len(ID),dtype=int)
for i in ax:
    sub1.plot([i,i],[ratio1_oxi[gtr][i] ,ratio2_oxi[gtr][i]] ,color="k")
    sub2.plot([i,i],[ratio1_H2O2[gtr][i],ratio2_H2O2[gtr][i]],color="k")
    
sub1.plot(ID,ratio1_oxi[gtr] ,linestyle="",marker="o",color="orange",label=r"$30min/0min$")
sub2.plot(ID,ratio1_H2O2[gtr],linestyle="",marker="o",color="orange",label=r"$60min/0min$")

sub1.plot(ID,ratio2_oxi[gtr] ,linestyle="",marker="o",color="r",label=r"$60min/0min$")
sub2.plot(ID,ratio2_H2O2[gtr],linestyle="",marker="o",color="r",label=r"$60min/0min$")

ymax = 80
sub1.set_ylim(1,ymax)
sub2.set_ylim(1,ymax)
sub1.legend()

sub1.tick_params("x",labelrotation=40)
sub2.tick_params("x",labelrotation=40)

sub1.set_title("oxi")
sub2.set_title("H2O2")


